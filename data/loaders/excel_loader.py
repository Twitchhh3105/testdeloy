"""Load .xlsx / .xls / .csv files into canonical sheet dicts."""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator


@dataclass
class LoadedSheet:
    source_type: str = "excel"
    file_name: str = ""
    sheet_name: str = ""
    headers: list[str] = field(default_factory=list)
    rows: list[list[str]] = field(default_factory=list)
    total_rows: int = 0


def _detect_header(values: list[list]) -> int:
    for i, row in enumerate(values):
        if any(cell not in (None, "") for cell in row):
            return i
    return 0


def _load_xlsx(path: Path) -> Iterator[LoadedSheet]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw = [list(row) for row in ws.iter_rows(values_only=True)]
        if not raw:
            continue
        header_idx = _detect_header(raw)
        headers = [
            (str(c) if c is not None else "") for c in raw[header_idx]
        ]
        body = raw[header_idx + 1 :]
        rows = [
            [("" if c is None else str(c)) for c in r]
            for r in body
            if any(cell not in (None, "") for cell in r)
        ]
        yield LoadedSheet(
            file_name=path.name,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            total_rows=len(rows),
        )


def _load_csv(path: Path) -> Iterator[LoadedSheet]:
    import pandas as pd

    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    yield LoadedSheet(
        file_name=path.name,
        sheet_name=path.stem,
        headers=[str(c) for c in df.columns],
        rows=df.astype(str).values.tolist(),
        total_rows=len(df),
    )


def load_excel_file(path: Path) -> Iterator[LoadedSheet]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls"}:
        yield from _load_xlsx(path)
    elif suffix == ".csv":
        yield from _load_csv(path)


def load_excel_files(root: Path) -> Iterator[LoadedSheet]:
    if not root.exists():
        return
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        if path.suffix.lower() not in {".xlsx", ".xls", ".csv"}:
            continue
        try:
            yield from load_excel_file(path)
        except Exception as exc:
            print(f"[excel_loader] failed on {path.name}: {exc}")
